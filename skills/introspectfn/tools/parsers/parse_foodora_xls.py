#!/usr/bin/env python3
"""Parse Foodora settlement XLS into normalized JSON.

Usage: parse_foodora_xls.py <path_to_xls>
Output: JSON to stdout
"""

import json
import os
import shutil
import sys
import tempfile

import openpyxl


def parse(file_path):
    # Foodora sends xlsx content with .xls extension — openpyxl needs .xlsx
    src = str(file_path)
    if src.lower().endswith(".xls"):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        shutil.copy2(src, tmp.name)
        tmp.close()
        wb = openpyxl.load_workbook(tmp.name)
        os.unlink(tmp.name)
    else:
        wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Read header row to find column positions
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            headers[cell.value.strip()] = col_idx

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Map by header position
        leverantor = row[headers.get("Leverantör", 4)] or ""
        if not leverantor.strip():
            continue  # skip daily aggregate rows

        order_id = row[headers.get("Order ID", 2)] or ""
        order_date = str(row[headers.get("Order Datum", 3)] or "")

        orders.append({
            "order_id": str(order_id),
            "date": order_date.split(" ")[0] if " " in order_date else order_date,
            "timestamp": order_date,
            "store_id": str(row[headers.get("Restaurang ID", 0)] or ""),
            "restaurant": str(row[headers.get("Kontonamn", 1)] or ""),
            "vat_0": float(row[headers.get("Underlag Moms 0%", 5)] or 0),
            "vat_6": float(row[headers.get("Moms 6%", 6)] or 0),
            "vat_12": float(row[headers.get("Moms 12%", 7)] or 0),
            "vat_25": float(row[headers.get("Moms 25%", 8)] or 0),
            "net_amount": float(row[headers.get("Nettobelopp", 9)] or 0),
            "gross_amount": float(row[headers.get("Bruttobelopp", 10)] or 0),
            "commission_base": float(row[headers.get("Provisionsgrundande belopp", 11)] or 0),
        })

    # Derive metadata from first order
    store_id = orders[0]["store_id"] if orders else ""
    restaurant = orders[0]["restaurant"] if orders else ""

    # Summary
    total_gross = sum(o["gross_amount"] for o in orders)
    total_net = sum(o["net_amount"] for o in orders)
    total_commission_base = sum(o["commission_base"] for o in orders)
    total_vat_6 = sum(o["vat_6"] for o in orders)
    total_vat_12 = sum(o["vat_12"] for o in orders)
    total_vat_25 = sum(o["vat_25"] for o in orders)

    return {
        "partner": "foodora",
        "store_id": store_id,
        "restaurant": restaurant,
        "orders": orders,
        "summary": {
            "total_orders": len(orders),
            "total_gross": round(total_gross, 2),
            "total_net": round(total_net, 2),
            "total_commission_base": round(total_commission_base, 2),
            "total_vat_6": round(total_vat_6, 2),
            "total_vat_12": round(total_vat_12, 2),
            "total_vat_25": round(total_vat_25, 2),
        },
    }


def main():
    if len(sys.argv) != 2:
        print(json.dumps({"error": "Usage: parse_foodora_xls.py <path>"}), file=sys.stderr)
        sys.exit(1)

    try:
        result = parse(sys.argv[1])
        print(json.dumps(result, indent=2, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
